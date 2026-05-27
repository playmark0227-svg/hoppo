#!/usr/bin/env python3
"""Build 北方領土ゲームアプリ クイズ全47問 一覧 (Excel)"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# --- Quiz data (mirrors js/data.js QUIZZES) ---
QUIZZES = [
    {
        "id": "basics",
        "title": "北方領土 基本クイズ",
        "color": "1F3D6E",  # navy
        "questions": [
            ("北方領土は、いくつの島でできている？",
             ["2つ", "3つ", "4つ", "5つ"], 2,
             "択捉島・国後島・色丹島・歯舞群島の4島をまとめて北方領土と呼ぶ。"),
            ("北方領土のうち、いちばん大きな島はどれ？",
             ["歯舞群島", "色丹島", "国後島", "択捉島"], 3,
             "択捉島（えとろふとう）は北方領土最大の島。本州の4分の1以上の大きさ。"),
            ("日本から北方領土に一番近い市はどこ？",
             ["稚内市", "根室市", "釧路市", "函館市"], 1,
             "根室市の納沙布岬から歯舞群島まで、わずか3.7km。"),
            ("「北方領土の日」は何月何日？",
             ["2月7日", "4月28日", "8月15日", "9月2日"], 0,
             "1855年2月7日、日魯通好条約が結ばれた日にちなんで制定。"),
            ("北方領土にかつて住んでいた日本人を何と呼ぶ？",
             ["海士", "島民", "元島民", "開拓民"], 2,
             "戦前に住み、終戦後に引き揚げた人たちを「元島民」と呼ぶ。"),
            ("北方四島の総面積はおよそどれくらい？",
             ["約500 km²", "約2,000 km²", "約5,000 km²", "約10,000 km²"], 2,
             "4島合わせて約5,036 km²。鳥取県より少し大きいくらい。"),
            ("戦前、北方領土には何人くらいの日本人が暮らしていた？",
             ["約3,000人", "約17,000人", "約50,000人", "約100,000人"], 1,
             "終戦時には約17,000人の日本人が4島で暮らしていた。"),
            ("北方領土問題で、日本が交渉している相手国は？",
             ["中国", "ロシア", "韓国", "アメリカ"], 1,
             "ソ連を引き継いだロシアと、平和条約の締結に向けて交渉が続いている。"),
            ("北方領土を管轄する日本の都道府県は？",
             ["青森県", "北海道", "岩手県", "秋田県"], 1,
             "北海道根室振興局の管轄。地図上では4島ともしっかり日本領として描かれている。"),
        ],
    },
    {
        "id": "history",
        "title": "歴史のクイズ",
        "color": "5E3A1F",  # brown
        "questions": [
            ("1855年に日本とロシアが結んだ、北方領土の国境を定めた条約は？",
             ["日魯通好条約", "樺太・千島交換条約", "ポーツマス条約", "サンフランシスコ平和条約"], 0,
             "日魯通好条約で、択捉島とウルップ島の間が国境と定められた。"),
            ("1875年、日本が千島列島全体をロシアから得るかわりに手放したのは？",
             ["北方四島", "樺太", "対馬", "千島列島南部"], 1,
             "樺太・千島交換条約で、日本は樺太（サハリン）を手放した。"),
            ("北方領土が不法に占拠されたのはいつ？",
             ["1904年", "1941年", "1945年", "1956年"], 2,
             "1945年8〜9月、終戦直後にソ連軍が4島を占領した。"),
            ("1956年に結ばれ、平和条約の締結後に2島返還を明記しているのは？",
             ["日ソ共同宣言", "ヤルタ協定", "日米安全保障条約", "東京宣言"], 0,
             "日ソ共同宣言では、平和条約後に歯舞群島と色丹島を返還すると明記。"),
            ("江戸時代後期、択捉島を探検し「大日本恵登呂府」の標柱を建てたのは？",
             ["伊能忠敬", "近藤重蔵", "間宮林蔵", "最上徳内"], 1,
             "1798年、近藤重蔵が標柱を建てて、ここが日本の領土であることを示した。"),
            ("1956年の日ソ共同宣言に署名した日本の首相は？",
             ["吉田茂", "鳩山一郎", "岸信介", "池田勇人"], 1,
             "当時の鳩山一郎首相がモスクワで署名。これで日ソの国交が回復した。"),
            ("ソ連崩壊後の1993年、日本とロシアが署名した宣言は？",
             ["東京宣言", "モスクワ宣言", "クラスノヤルスク合意", "シベリア宣言"], 0,
             "東京宣言では、4島の帰属問題を解決して平和条約を結ぶ方針が確認された。"),
            ("「北方領土の日」が制定されたのは何年？",
             ["1956年", "1981年", "1991年", "2001年"], 1,
             "1981年、閣議決定で2月7日を「北方領土の日」と定めた。"),
        ],
    },
    {
        "id": "nature",
        "title": "自然・文化クイズ",
        "color": "1F6E4A",  # green
        "questions": [
            ("北方領土の海によくいる、かわいい哺乳類は？",
             ["アザラシ", "パンダ", "コアラ", "カピバラ"], 0,
             "ゴマフアザラシなど、たくさんのアザラシが生息している。"),
            ("流氷が接岸する海は？",
             ["日本海", "瀬戸内海", "オホーツク海", "太平洋"], 2,
             "オホーツク海の流氷は世界でも低緯度で見られる珍しいもの。"),
            ("根室名物で、北方領土の海でも獲れるカニは？",
             ["タラバガニ", "ズワイガニ", "花咲ガニ", "毛ガニ"], 2,
             "花咲ガニは根室のシンボル。濃厚な味わいで知られる。"),
            ("オホーツク海にいる、「流氷の天使」と呼ばれる生き物は？",
             ["クリオネ", "クラゲ", "イルカ", "シャチ"], 0,
             "クリオネは小さな巻貝のなかま。妖精のようにひらひら泳ぐ。"),
            ("北方領土最高峰の山は、どの島にある？",
             ["択捉島", "国後島", "色丹島", "歯舞群島"], 1,
             "国後島の爺爺岳（ちゃちゃだけ）は1,822m。北方領土でいちばん高い。"),
            ("国後島のシンボルとなっている美しい火山の名前は？",
             ["羅臼岳", "爺爺岳", "泊山", "斜古丹山"], 1,
             "爺爺岳（ちゃちゃだけ）は富士山に似た円錐形の活火山。"),
            ("北方領土の周辺の海で、戦前にたくさん獲れた水産物は？",
             ["マグロ・カツオ", "鮭・マス・昆布", "タイ・ヒラメ", "ウナギ・アユ"], 1,
             "鮭・マス・ニシン漁や昆布漁が盛んで、缶詰工場もあった。"),
            ("北方領土で見られる、大きな哺乳類の代表は？",
             ["ヒグマ", "パンダ", "ライオン", "チンパンジー"], 0,
             "択捉島や国後島にはヒグマが生息していて、サケを捕まえる姿も見られる。"),
        ],
    },
    {
        "id": "islands",
        "title": "四島それぞれクイズ",
        "color": "5A4E8C",  # purple
        "questions": [
            ("国後島の面積はおよそどれくらい？",
             ["約250 km²", "約1,500 km²", "約3,000 km²", "約5,000 km²"], 1,
             "国後島は約1,489 km²。沖縄本島より大きい。"),
            ("色丹島の最高峰の山は？",
             ["爺爺岳", "茂世路岳", "斜古丹山", "羅臼岳"], 2,
             "斜古丹山（しゃこたんやま）412m。なだらかな丘陵が広がる島。"),
            ("歯舞群島の中で、納沙布岬から最も近い島は？",
             ["貝殻島", "志発島", "勇留島", "水晶島"], 0,
             "貝殻島はわずか3.7kmの距離。岬から灯台が肉眼で見える。"),
            ("国後島と北海道（根室半島）の最短距離はおよそ？",
             ["約3 km", "約16 km", "約50 km", "約100 km"], 1,
             "根室半島と国後島はわずか16km。晴れた日には島影がはっきり見える。"),
            ("戦前の択捉島で中心地として栄えていた町は？",
             ["内岡（紗那）", "穴澗", "古釜布", "泊"], 0,
             "内岡（うちおか）の紗那（しゃな）地区が行政・経済の中心だった。"),
            ("歯舞群島は何でできた島々の集まり？",
             ["ひとつの大きな島", "6つの島と岩礁", "20以上の島々", "無人島だけ"], 1,
             "貝殻島・水晶島・秋勇留島・勇留島・志発島・多楽島など、6島群と岩礁から成る。"),
        ],
    },
    {
        "id": "expert",
        "title": "難問チャレンジ",
        "color": "8C2E2E",  # deep red
        "questions": [
            ("1855年の日魯通好条約は、現在の何という場所で調印された？",
             ["江戸（東京）", "下田", "函館", "モスクワ"], 1,
             "伊豆半島の下田、玉泉寺で調印。プチャーチン提督と川路聖謨らが交渉した。"),
            ("近藤重蔵が1798年に択捉島に建てた標柱に書かれていた言葉は？",
             ["大日本恵登呂府", "日本国之地", "皇国之島", "蝦夷地以北"], 0,
             "「大日本恵登呂府（だいにっぽんえとろふ）」と書かれ、日本領であることを示した。"),
            ("1875年の樺太・千島交換条約で日本領となった、千島列島最北端の島は？",
             ["ウルップ島", "パラムシル島", "シュムシュ島（占守島）", "ラショワ島"], 2,
             "シュムシュ島はカムチャッカ半島のすぐ南。樺太を手放す代わりに千島列島全体を得た。"),
            ("1945年2月のヤルタ会談で千島列島の引き渡しを密約したのは？",
             ["米ソの2カ国", "米英ソの3カ国", "米英中ソの4カ国", "英ソの2カ国"], 1,
             "ルーズベルト・チャーチル・スターリンによる米英ソの密約。後に大きな問題となった。"),
            ("択捉島と国後島の間を流れる海峡の名前は？",
             ["根室海峡", "国後水道", "択捉水道", "宗谷海峡"], 1,
             "国後水道（くなしりすいどう）と呼ばれる海峡。"),
            ("江戸時代に「赤蝦夷風説考」を著し、ロシアの脅威を警告した人物は？",
             ["林子平", "工藤平助", "本多利明", "大黒屋光太夫"], 1,
             "仙台藩医・工藤平助が1781年に田沼意次に提出。日本初のロシア研究書とも言われる。"),
            ("1812年、ロシアに捕らえられたゴロウニン艦長の解放交渉に尽力した日本の商人は？",
             ["大黒屋光太夫", "高田屋嘉兵衛", "津太夫", "中川五郎治"], 1,
             "淡路出身の高田屋嘉兵衛は、自身もロシアに連行された後、両国の橋渡し役となった。"),
            ("北方領土を「南クリル」と呼ぶロシアの管轄行政区分は？",
             ["カムチャッカ地方", "マガダン州", "サハリン州", "沿海地方"], 2,
             "ロシアでは南クリル管区としてサハリン州の管轄下に置かれている。"),
            ("北方領土最高峰・爺爺岳（ちゃちゃだけ）の正確な標高は？",
             ["1,498m", "1,772m", "1,822m", "1,917m"], 2,
             "国後島の爺爺岳は1,822m。富士山に似た円錐形の活火山。"),
            ("根室市にある北方領土に関する学習・展示施設「ニ・ホ・ロ」とは？",
             ["北方四島交流センター", "北方領土歴史博物館", "北方資料館", "オホーツク文化館"], 0,
             "正式名「北方四島交流センター」。1986年に根室市にオープンした学習・交流拠点。"),
        ],
    },
    {
        "id": "lifestyle",
        "title": "島の暮らしクイズ",
        "color": "8C5E1F",  # warm brown
        "questions": [
            ("戦前、北方領土には町や村がいくつあった？",
             ["7つ", "15", "24", "40以上"], 2,
             "4島合わせて24の町や村があり、学校や病院、役場もそろっていた。"),
            ("戦前の北方領土で日本人が多く従事していた仕事は？",
             ["工場労働", "漁業・水産加工", "農業", "林業"], 1,
             "鮭・マス・昆布の漁業と、それを加工する缶詰工場が島の経済を支えた。"),
            ("元島民が島を訪れる「ビザなし渡航」が始まったのは何年？",
             ["1956年", "1965年", "1992年", "2005年"], 2,
             "1992年から始まった人道的な交流。元島民がパスポートなしで島を訪問できた。"),
            ("戦後、元島民が初めて先祖の墓参りに島を訪れたのは何年？",
             ["1956年", "1964年", "1972年", "1985年"], 1,
             "1964年、ようやく墓参りが認められて、元島民の悲願がひとつ叶った。"),
            ("元島民やその子孫が今いちばん多く暮らす街は？",
             ["札幌市", "函館市", "根室市", "釧路市"], 2,
             "根室市は北方領土を望む最前線の街。返還運動の中心地でもある。"),
            ("北方領土返還運動でよく使われる合言葉は？",
             ["返せ！北方領土", "取り戻せ！日本の島", "帰ろう！故郷へ", "守ろう！四島"], 0,
             "「返せ！北方領土」は返還運動を象徴する合言葉として広く知られている。"),
        ],
    },
]

# --- Styles ---
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(name="Yu Gothic", size=18, bold=True, color="1F3D6E")
header_font = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F3D6E")
cat_font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
correct_fill = PatternFill("solid", fgColor="DFF0DC")  # light green
correct_font = Font(name="Yu Gothic", size=10, bold=True, color="1F7D3E")
choice_font = Font(name="Yu Gothic", size=10, color="2A2A2A")
body_font = Font(name="Yu Gothic", size=10, color="2A2A2A")
mincho_body_font = Font(name="Yu Mincho", size=10, color="2A2A2A")
note_font = Font(name="Yu Mincho", size=9, italic=True, color="666666")
alt_row_fill = PatternFill("solid", fgColor="F8F9FB")

# --- Build workbook ---
wb = openpyxl.Workbook()

# === Sheet 1: 全問一覧 ===
ws = wb.active
ws.title = "全問一覧"

# Row 1: document title
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "北方領土ゲームアプリ クイズ 全47問 一覧"
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

# Row 2: subtitle / meta
ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = "出典：内閣府・北方領土問題対策協会・外務省 公表資料、自治体公表資料、学術文献　／　全問正解時はカテゴリごとに +50pt のボーナス　／　定期的に内容を見直し・更新します"
c.font = note_font
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

# Row 3: blank spacer
ws.row_dimensions[3].height = 6

# Row 4: header
HEADERS = ["No.", "カテゴリ", "通番", "設問", "選択肢1", "選択肢2", "選択肢3", "選択肢4", "正解", "解説"]
# Adjust HEADERS — A No., B Cat, C 通番, D 設問, E-H 選択肢, I 正解, J 解説
HEADERS = ["No.", "カテゴリ", "カテゴリ内番号", "設問", "選択肢 1", "選択肢 2", "選択肢 3", "選択肢 4", "正解", "解説"]
for i, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=4, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[4].height = 30

# Body rows
row = 5
n_total = 0
for cat in QUIZZES:
    cat_start_row = row
    cat_title = cat["title"]
    cat_color = cat["color"]
    cat_fill = PatternFill("solid", fgColor=cat_color)
    for idx, (q, choices, ans, expl) in enumerate(cat["questions"], start=1):
        n_total += 1
        # Pad choices to 4
        while len(choices) < 4:
            choices.append("")
        values = [
            n_total,
            cat_title,
            f"{idx}/{len(cat['questions'])}",
            q,
            choices[0], choices[1], choices[2], choices[3],
            f"{['A','B','C','D'][ans]}：{choices[ans]}",
            expl,
        ]
        for col_i, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_i, value=v)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_i in (1, 3) else "left",
                vertical="top" if col_i in (4, 10) else "center",
                wrap_text=True,
            )
            if col_i == 2:
                # category column — colored
                cell.fill = cat_fill
                cell.font = cat_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_i in (5, 6, 7, 8):
                cell.font = choice_font
                if col_i - 5 == ans:
                    cell.fill = correct_fill
                    cell.font = correct_font
            elif col_i == 9:
                cell.font = correct_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_i == 10:
                cell.font = mincho_body_font
            else:
                cell.font = body_font
        # Set per-row height based on question length (heuristic)
        q_lines = max(2, (len(q) // 24) + 1)
        e_lines = max(2, (len(expl) // 28) + 1)
        max_lines = max(q_lines, e_lines, 2)
        ws.row_dimensions[row].height = max(38, 16 * max_lines)
        row += 1

# Column widths
col_widths = {
    "A": 5,    # No.
    "B": 18,   # カテゴリ
    "C": 10,   # カテゴリ内番号
    "D": 36,   # 設問
    "E": 18, "F": 18, "G": 18, "H": 18,  # 選択肢 1-4
    "I": 16,   # 正解
    "J": 44,   # 解説
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Freeze top header
ws.freeze_panes = "A5"

# === Sheet 2: カテゴリ別 (sub-sheets) ===
for cat in QUIZZES:
    ws2 = wb.create_sheet(title=cat["title"][:31])
    cat_fill = PatternFill("solid", fgColor=cat["color"])

    # Title
    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = cat["title"]
    c.font = Font(name="Yu Gothic", size=16, bold=True, color="FFFFFF")
    c.fill = cat_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    # Header
    headers = ["No.", "設問", "選択肢 1", "選択肢 2", "選択肢 3", "選択肢 4", "正解", "解説"]
    for i, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws2.row_dimensions[2].height = 28

    # Adjust to have 8 columns: No, q, 4 choices, ans, exp
    for i in range(len(headers) + 1, 9):  # ensure A..H
        pass

    # Re-do header with 8 columns
    ws2.merge_cells("A1:H1")
    r = 3
    for idx, (q, choices, ans, expl) in enumerate(cat["questions"], start=1):
        while len(choices) < 4:
            choices.append("")
        values = [
            idx,
            q,
            choices[0], choices[1], choices[2], choices[3],
            f"{['A','B','C','D'][ans]}：{choices[ans]}",
            expl,
        ]
        for col_i, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=col_i, value=v)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_i in (1, 7) else "left",
                vertical="top" if col_i in (2, 8) else "center",
                wrap_text=True,
            )
            if col_i in (3, 4, 5, 6):
                cell.font = choice_font
                if col_i - 3 == ans:
                    cell.fill = correct_fill
                    cell.font = correct_font
            elif col_i == 7:
                cell.font = correct_font
            elif col_i == 8:
                cell.font = mincho_body_font
            else:
                cell.font = body_font
        q_lines = max(2, (len(q) // 24) + 1)
        e_lines = max(2, (len(expl) // 28) + 1)
        max_lines = max(q_lines, e_lines, 2)
        ws2.row_dimensions[r].height = max(38, 16 * max_lines)
        r += 1

    # Widths
    widths = {"A": 5, "B": 40, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 44}
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A3"

# === Sheet 3: 概要 ===
ws3 = wb.create_sheet(title="概要", index=0)
ws3.merge_cells("A1:D1")
c = ws3["A1"]
c.value = "北方領土ゲームアプリ クイズ 一覧（概要）"
c.font = title_font
c.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 30

ws3["A3"] = "アプリ名"; ws3["B3"] = "北方領土ステータス"
ws3["A4"] = "URL";     ws3["B4"] = "https://playmark0227-svg.github.io/hoppo/"
ws3["A5"] = "作成日";   ws3["B5"] = "2026年5月24日"
ws3["A6"] = "版";       ws3["B6"] = "第 1 版"
ws3["A7"] = "備考";     ws3["B7"] = "内容は定期的に見直し・更新を行います。"

for row_i in range(3, 8):
    for col_i in range(1, 3):
        cell = ws3.cell(row=row_i, column=col_i)
        cell.font = Font(name="Yu Gothic", size=11, bold=(col_i == 1), color="2A2A2A")
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Category summary table
hdr_row = 9
headers3 = ["No.", "カテゴリ", "設問数", "備考"]
for i, h in enumerate(headers3, start=1):
    cell = ws3.cell(row=hdr_row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
ws3.row_dimensions[hdr_row].height = 26

r = hdr_row + 1
total_q = 0
for idx, cat in enumerate(QUIZZES, start=1):
    qn = len(cat["questions"])
    total_q += qn
    cat_fill = PatternFill("solid", fgColor=cat["color"])
    values = [idx, cat["title"], f"{qn} 問", "全問正解で +50pt"]
    for col_i, v in enumerate(values, start=1):
        cell = ws3.cell(row=r, column=col_i, value=v)
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col_i in (1, 3) else "left", vertical="center")
        cell.font = body_font
        if col_i == 2:
            cell.fill = cat_fill
            cell.font = cat_font
    ws3.row_dimensions[r].height = 24
    r += 1

# Total row
cell = ws3.cell(row=r, column=1, value="合計")
cell.font = Font(name="Yu Gothic", size=11, bold=True, color="2A2A2A")
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = PatternFill("solid", fgColor="F2F4F8")
cell.border = border
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
cell = ws3.cell(row=r, column=2)
cell.alignment = Alignment(horizontal="right", vertical="center")
cell = ws3.cell(row=r, column=3, value=f"{total_q} 問")
cell.font = Font(name="Yu Gothic", size=11, bold=True, color="1F3D6E")
cell.fill = PatternFill("solid", fgColor="F2F4F8")
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = border
cell = ws3.cell(row=r, column=4, value="")
cell.fill = PatternFill("solid", fgColor="F2F4F8")
cell.border = border
ws3.row_dimensions[r].height = 26

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 28

# Save
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "北方領土ゲームアプリ_クイズ全47問_v1.xlsx")
wb.save(out_path)
print("Created:", out_path)
print(f"  Total questions: {total_q}")
print(f"  Sheets: {wb.sheetnames}")
